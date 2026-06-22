"""
Monthly investor report — Excel (.xlsx) and PDF exports.
Excel: 5 sheets — Summary, Holdings, Income Calendar, Dividend Events, Received Dividends
PDF:  printable A4 report with cover metrics, holdings table, calendar, received dividends
"""
import io
from datetime import date, timedelta, datetime

from fastapi import APIRouter
from fastapi.responses import StreamingResponse

from database import db
from services.dividend_fetcher import detect_frequency, project_future_payments

router = APIRouter(prefix="/api/export", tags=["export"])


# ── Shared data fetch ──────────────────────────────────────────────────────────

def _load_report_data() -> dict:
    with db() as conn:
        positions_raw = [dict(r) for r in conn.execute("SELECT * FROM positions ORDER BY name").fetchall()]
        events_raw = [dict(r) for r in conn.execute(
            """SELECT de.*, p.name as position_name, p.units
               FROM dividend_events de JOIN positions p ON p.id=de.position_id
               ORDER BY de.ex_date DESC"""
        ).fetchall()]
        received_raw = [dict(r) for r in conn.execute(
            """SELECT rd.*, p.name as position_name
               FROM received_dividends rd JOIN positions p ON p.id=rd.position_id
               ORDER BY rd.pay_date DESC"""
        ).fetchall()]
        cutoff = (date.today() - timedelta(days=365)).isoformat()
        row = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as t FROM received_dividends WHERE pay_date >= ?", (cutoff,)
        ).fetchone()
        try:
            cash_row = conn.execute(
                "SELECT value FROM portfolio_settings WHERE key='cash_balance'"
            ).fetchone()
            cash_balance = float(cash_row["value"]) if cash_row else 0.0
        except Exception:
            cash_balance = 0.0

    income_ttm = float(row["t"]) if row else 0.0

    for p in positions_raw:
        units = p.get("units", 0)
        book = p.get("book_cost_per_unit", 0)
        tb = round(units * book, 2)
        p["total_book_cost"] = tb
        px = p.get("last_price")
        if px is not None:
            cv = round(units * px, 2)
            p["current_value"] = cv
            p["unrealised_pnl"] = round(cv - tb, 2)
            p["unrealised_pnl_pct"] = round((cv - tb) / tb * 100, 2) if tb else 0.0
        else:
            p["current_value"] = p["unrealised_pnl"] = p["unrealised_pnl_pct"] = None

    total_book = sum(p["total_book_cost"] for p in positions_raw)
    total_value = sum(p["current_value"] for p in positions_raw if p["current_value"] is not None)
    capital_growth = round(total_value - total_book, 2)
    capital_growth_pct = round(capital_growth / total_book * 100, 2) if total_book else 0.0

    events_by_pos: dict = {}
    for e in events_raw:
        events_by_pos.setdefault(e["position_id"], []).append(e)

    calendar: dict = {}
    for pos in positions_raw:
        evs = events_by_pos.get(pos["id"], [])
        if not evs:
            continue
        freq = detect_frequency(evs)
        for pmt in project_future_payments(pos, evs, freq):
            calendar.setdefault(pmt["year_month"], []).append(pmt)

    proj_total = sum(
        pmt["projected_total"]
        for months in calendar.values()
        for pmt in months
    )

    by_asset: dict = {}
    for p in positions_raw:
        at = p.get("asset_type", "stock")
        by_asset.setdefault(at, {"book_cost": 0.0, "value": 0.0, "count": 0})
        by_asset[at]["book_cost"] += p["total_book_cost"]
        by_asset[at]["count"] += 1
        if p["current_value"] is not None:
            by_asset[at]["value"] += p["current_value"]

    return {
        "report_date": date.today().strftime("%B %Y"),
        "iso_date": date.today().isoformat(),
        "positions": positions_raw,
        "events": events_raw,
        "events_by_pos": events_by_pos,
        "received": received_raw,
        "calendar": calendar,
        "summary": {
            "total_book_cost": round(total_book, 2),
            "total_current_value": round(total_value, 2),
            "cash_balance": round(cash_balance, 2),
            "total_with_cash": round(total_value + cash_balance, 2),
            "capital_growth": capital_growth,
            "capital_growth_pct": capital_growth_pct,
            "income_ttm": round(income_ttm, 2),
            "projected_annual_income": round(proj_total, 2),
            "positions_count": len(positions_raw),
            "by_asset_type": {k: {
                "book_cost": round(v["book_cost"], 2),
                "value": round(v["value"], 2),
                "count": v["count"],
            } for k, v in by_asset.items()},
        },
    }


# ════════════════════════════════════════════════════════════════════════════════
# EXCEL
# ════════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BG_DARK  = "0D1117"; _BG_SURF = "161B22"; _BG_SURF2 = "21262D"
_ACCENT   = "388BFD"; _POS     = "3FB950"; _NEG      = "F85149"
_MUTED    = "8B949E"; _TEXT    = "E6EDF3"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(h=_TEXT, bold=False, size=10): return Font(color=h, bold=bold, size=size, name="Calibri")
def _bdr():
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, cols):
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=lbl)
        cell.fill = _fill(_BG_SURF2); cell.font = _font(_MUTED, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = _bdr()

def _dc(ws, row, col, val, fmt=None, color=_TEXT, bold=False, align="left"):
    cell = ws.cell(row=row, column=col, value=val)
    cell.fill = _fill(_BG_SURF if row % 2 == 0 else _BG_DARK)
    cell.font = _font(color, bold); cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _bdr()
    if fmt and val is not None: cell.number_format = fmt

def _autowidth(ws, mn=10, mx=45):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(mn, min(w + 2, mx))


def _xl_summary(wb, summary, report_date):
    ws = wb.create_sheet("Summary"); ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _ACCENT
    ws.merge_cells("A1:D1")
    t = ws["A1"]; t.value = f"Portfolio Monthly Report — {report_date}"
    t.fill = _fill(_BG_SURF); t.font = _font(_TEXT, bold=True, size=14)
    t.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:D2"); ws["A2"].fill = _fill(_BG_DARK)

    cg = summary.get("capital_growth") or 0
    cash = summary.get("cash_balance") or 0
    metrics = [
        ("Invested Portfolio Value",     summary.get("total_current_value"),     "£#,##0.00", _ACCENT, True),
        ("Cash Balance",                 cash,                                    "£#,##0.00", _TEXT,   False),
        ("Total Portfolio (incl. cash)", summary.get("total_with_cash"),          "£#,##0.00", _ACCENT, True),
        ("Total Book Cost",              summary.get("total_book_cost"),          "£#,##0.00", _TEXT,   False),
        ("Capital Growth £",             cg,                                      "£#,##0.00", _POS if cg >= 0 else _NEG, True),
        ("Capital Growth %",             summary.get("capital_growth_pct"),       '0.00"%"',   _POS if cg >= 0 else _NEG, False),
        ("",                             None, None, _TEXT, False),
        ("Income Received (12m)",        summary.get("income_ttm"),               "£#,##0.00", _POS, True),
        ("Projected Annual Income",      summary.get("projected_annual_income"),  "£#,##0.00", _POS, False),
        ("",                             None, None, _TEXT, False),
        ("Total Positions",              summary.get("positions_count"),          "0",          _TEXT, False),
    ]
    for i, (lbl, val, fmt, color, bold) in enumerate(metrics, start=3):
        lc = ws.cell(row=i, column=1, value=lbl)
        bg = _BG_SURF if i % 2 == 0 else _BG_DARK
        lc.fill = _fill(bg); lc.font = _font(_MUTED); lc.border = _bdr()
        lc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{i}:C{i}")
        vc = ws.cell(row=i, column=4, value=val)
        vc.fill = _fill(bg); vc.font = _font(color, bold)
        vc.alignment = Alignment(horizontal="right", vertical="center"); vc.border = _bdr()
        if fmt and val is not None: vc.number_format = fmt

    r = len(metrics) + 4
    ws.cell(row=r, column=1, value="Asset Class Breakdown").font = _font(_TEXT, bold=True); r += 1
    _hdr(ws, r, ["Asset Class", "Positions", "Book Cost £", "Market Value £"]); r += 1
    for at, d in (summary.get("by_asset_type") or {}).items():
        _dc(ws, r, 1, at.replace("_"," ").title())
        _dc(ws, r, 2, d["count"], align="center")
        _dc(ws, r, 3, d["book_cost"], fmt="£#,##0.00", align="right")
        _dc(ws, r, 4, d["value"], fmt="£#,##0.00", align="right", color=_ACCENT); r += 1

    for col, w in zip("ABCD", [28, 28, 28, 18]):
        ws.column_dimensions[col].width = w


def _xl_holdings(wb, positions):
    ws = wb.create_sheet("Holdings"); ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _ACCENT; ws.freeze_panes = "A2"
    cols = ["Name","ISIN","Ticker","Type","Units","Book/Unit £","Total Cost £",
            "Native Price","CCY","GBP Price £","Value £","PnL £","PnL %",
            "Div Yield %","FX Rate","Source","Price Updated"]
    _hdr(ws, 1, cols)
    for r, p in enumerate(positions, start=2):
        pnl = p.get("unrealised_pnl"); pc = _POS if (pnl or 0) >= 0 else _NEG
        dy = p.get("annual_yield")
        _dc(ws,r,1, p.get("name"), bold=True)
        _dc(ws,r,2, p.get("isin") or "")
        _dc(ws,r,3, p.get("ticker") or "")
        _dc(ws,r,4, (p.get("asset_type") or "").replace("_"," ").title())
        _dc(ws,r,5, p.get("units"),                fmt="#,##0.######",  align="right")
        _dc(ws,r,6, p.get("book_cost_per_unit"),   fmt="£#,##0.0000",   align="right")
        _dc(ws,r,7, p.get("total_book_cost"),      fmt="£#,##0.00",     align="right")
        _dc(ws,r,8, p.get("native_price"),         fmt="#,##0.0000",    align="right")
        _dc(ws,r,9, p.get("native_currency") or "")
        _dc(ws,r,10,p.get("last_price"),           fmt="£#,##0.0000",   align="right", color=_ACCENT)
        _dc(ws,r,11,p.get("current_value"),        fmt="£#,##0.00",     align="right", color=_ACCENT)
        _dc(ws,r,12,pnl,                           fmt="£#,##0.00",     align="right", color=pc)
        _dc(ws,r,13,p.get("unrealised_pnl_pct"),   fmt='0.00"%"',       align="right", color=pc)
        _dc(ws,r,14,(dy*100) if dy is not None else None, fmt='0.00"%"',align="right", color=_POS if dy else _TEXT)
        _dc(ws,r,15,p.get("last_fx_rate"),         fmt="0.0000",        align="right")
        _dc(ws,r,16,p.get("last_price_source") or "")
        _dc(ws,r,17,p.get("last_price_at") or "")
    _autowidth(ws)


def _xl_calendar(wb, calendar):
    ws = wb.create_sheet("Income Calendar"); ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _POS
    _hdr(ws, 1, ["Month","Position","Frequency","Projected Date","Projected Total £"])
    r = 2; grand = 0.0
    for ym in sorted(calendar):
        lbl = datetime.strptime(ym+"-01","%Y-%m-%d").strftime("%B %Y")
        mtotal = sum(p["projected_total"] for p in calendar[ym]); grand += mtotal
        for c in range(1,6):
            cell = ws.cell(row=r,column=c); cell.fill=_fill(_BG_SURF2); cell.border=_bdr()
        ws.cell(row=r,column=1,value=lbl).font=_font(_TEXT,bold=True)
        ws.cell(row=r,column=1).fill=_fill(_BG_SURF2)
        tc=ws.cell(row=r,column=5,value=mtotal); tc.font=_font(_POS,bold=True)
        tc.number_format="£#,##0.00"; tc.alignment=Alignment(horizontal="right"); tc.fill=_fill(_BG_SURF2); r+=1
        for p in sorted(calendar[ym],key=lambda x:-x["projected_total"]):
            _dc(ws,r,1,""); _dc(ws,r,2,p["position_name"]); _dc(ws,r,3,p["frequency_label"])
            _dc(ws,r,4,p["projected_date"]); _dc(ws,r,5,p["projected_total"],fmt="£#,##0.00",align="right",color=_POS); r+=1
    for c in range(1,6):
        cell=ws.cell(row=r,column=c); cell.fill=_fill(_BG_SURF); cell.border=_bdr()
    ws.cell(row=r,column=1,value="12-Month Total").font=_font(_TEXT,bold=True)
    ws.cell(row=r,column=1).fill=_fill(_BG_SURF)
    gc=ws.cell(row=r,column=5,value=grand); gc.font=_font(_POS,bold=True,size=11)
    gc.number_format="£#,##0.00"; gc.alignment=Alignment(horizontal="right"); gc.fill=_fill(_BG_SURF)
    _autowidth(ws)


def _xl_events(wb, events):
    ws = wb.create_sheet("Dividend Events"); ws.sheet_view.showGridLines = False; ws.freeze_panes="A2"
    _hdr(ws,1,["Position","Ex-Date","Pay Date","Amount/Unit","Currency","Type","Source"])
    for r,e in enumerate(events,start=2):
        _dc(ws,r,1,e.get("position_name"),bold=True); _dc(ws,r,2,e.get("ex_date"))
        _dc(ws,r,3,e.get("pay_date") or ""); _dc(ws,r,4,e.get("amount_per_unit"),fmt="#,##0.000000",align="right",color=_POS)
        _dc(ws,r,5,e.get("currency") or ""); _dc(ws,r,6,e.get("div_type") or ""); _dc(ws,r,7,e.get("source") or "")
    _autowidth(ws)


def _xl_received(wb, received):
    ws = wb.create_sheet("Received Dividends"); ws.sheet_view.showGridLines = False; ws.freeze_panes="A2"
    _hdr(ws,1,["Position","Pay Date","Amount £","Currency","Notes"])
    total = 0.0
    for r,d in enumerate(received,start=2):
        _dc(ws,r,1,d.get("position_name"),bold=True); _dc(ws,r,2,d.get("pay_date"))
        _dc(ws,r,3,d.get("amount"),fmt="£#,##0.00",align="right",color=_POS)
        _dc(ws,r,4,d.get("currency") or ""); _dc(ws,r,5,d.get("notes") or ""); total += d.get("amount") or 0
    r = len(received)+2
    for c in range(1,6):
        cell=ws.cell(row=r,column=c); cell.fill=_fill(_BG_SURF); cell.border=_bdr()
    ws.cell(row=r,column=1,value="Total Received").font=_font(_TEXT,bold=True)
    ws.cell(row=r,column=1).fill=_fill(_BG_SURF)
    tc=ws.cell(row=r,column=3,value=total); tc.font=_font(_POS,bold=True)
    tc.number_format="£#,##0.00"; tc.alignment=Alignment(horizontal="right"); tc.fill=_fill(_BG_SURF)
    _autowidth(ws)


@router.get("/report.xlsx")
def export_xlsx():
    d = _load_report_data()
    wb = Workbook(); wb.remove(wb.active)
    _xl_summary(wb, d["summary"], d["report_date"])
    _xl_holdings(wb, d["positions"])
    _xl_calendar(wb, d["calendar"])
    _xl_events(wb, d["events"])
    _xl_received(wb, d["received"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"portfolio_report_{d['iso_date']}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ════════════════════════════════════════════════════════════════════════════════
# PDF
# ════════════════════════════════════════════════════════════════════════════════

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, HRFlowable
)


_PDF_BG      = colors.HexColor("#0D1117")
_PDF_SURF    = colors.HexColor("#161B22")
_PDF_SURF2   = colors.HexColor("#21262D")
_PDF_ACCENT  = colors.HexColor("#388BFD")
_PDF_POS     = colors.HexColor("#3FB950")
_PDF_NEG     = colors.HexColor("#F85149")
_PDF_MUTED   = colors.HexColor("#8B949E")
_PDF_TEXT    = colors.HexColor("#E6EDF3")
_PDF_BORDER  = colors.HexColor("#30363D")

def _gbp(v): return f"£{v:,.2f}" if v is not None else "—"
def _pct(v): return f"{v:+.2f}%" if v is not None else "—"
def _num(v, dp=4): return f"{v:,.{dp}f}" if v is not None else "—"

def _pdf_table_style(header_rows=1):
    return TableStyle([
        ("BACKGROUND",   (0, 0),           (-1, header_rows-1), _PDF_SURF2),
        ("TEXTCOLOR",    (0, 0),           (-1, header_rows-1), _PDF_MUTED),
        ("FONTNAME",     (0, 0),           (-1, header_rows-1), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0),           (-1, header_rows-1), 7),
        ("ALIGN",        (0, 0),           (-1, header_rows-1), "CENTER"),
        ("BACKGROUND",   (0, header_rows), (-1, -1),            _PDF_SURF),
        ("TEXTCOLOR",    (0, header_rows), (-1, -1),            _PDF_TEXT),
        ("FONTNAME",     (0, header_rows), (-1, -1),            "Helvetica"),
        ("FONTSIZE",     (0, header_rows), (-1, -1),            7),
        ("TOPPADDING",   (0, 0),           (-1, -1),            3),
        ("BOTTOMPADDING",(0, 0),           (-1, -1),            3),
        ("LEFTPADDING",  (0, 0),           (-1, -1),            5),
        ("RIGHTPADDING", (0, 0),           (-1, -1),            5),
        ("GRID",         (0, 0),           (-1, -1),            0.3, _PDF_BORDER),
        ("VALIGN",       (0, 0),           (-1, -1),            "MIDDLE"),
    ])


import httpx
from routers.ai import get_ollama_settings

async def _fetch_ai_summary(d: dict) -> str:
    settings = get_ollama_settings()
    s = d["summary"]
    prompt = f"""You are a professional financial analyst. Write a concise 1-2 paragraph executive summary for a monthly portfolio report.
The portfolio has £{s['total_current_value']} in assets across {s['positions_count']} positions.
It has generated £{s['capital_growth']} in capital growth ({s['capital_growth_pct']}%) and is projected to yield £{s['projected_annual_income']} in dividends over the next 12 months.
Do not use markdown. Write in a formal, encouraging tone."""
    
    ollama_url = settings["url"].rstrip('/') + "/api/generate"
    payload = {"model": settings["model"], "prompt": prompt, "stream": False}
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(ollama_url, json=payload)
            if resp.status_code == 200:
                data = resp.json()
                return data.get("response", "").strip()
    except Exception:
        pass
    return "AI-generated executive summary is currently unavailable. Please ensure your local Ollama instance is running."

async def _build_pdf_async(d: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)

    h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=20,
                         textColor=_PDF_TEXT, spaceAfter=4, backColor=_PDF_BG)
    h2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=12,
                         textColor=_PDF_ACCENT, spaceAfter=6, spaceBefore=12)
    h3 = ParagraphStyle("h3", fontName="Helvetica-Bold", fontSize=9,
                         textColor=_PDF_MUTED, spaceAfter=4, spaceBefore=8)
    sub = ParagraphStyle("sub", fontName="Helvetica", fontSize=9,
                          textColor=_PDF_MUTED, spaceAfter=12)
    body_text = ParagraphStyle("body_text", fontName="Helvetica", fontSize=9,
                                textColor=_PDF_TEXT, spaceAfter=12, leading=14)

    s = d["summary"]
    story = []

    # ── Cover ──────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph("Portfolio Monthly Report", h1))
    story.append(Paragraph(d["report_date"], sub))
    story.append(HRFlowable(width="100%", thickness=1, color=_PDF_ACCENT, spaceAfter=8))

    cg = s.get("capital_growth") or 0
    cash = s.get("cash_balance") or 0
    cover_data = [
        ["Invested Portfolio Value",  _gbp(s.get("total_current_value"))],
        ["Cash Balance",              _gbp(cash)],
        ["Total Portfolio (incl. cash)", _gbp(s.get("total_with_cash"))],
        ["Total Book Cost",           _gbp(s.get("total_book_cost"))],
        ["Capital Growth",            f"{_gbp(cg)}  ({_pct(s.get('capital_growth_pct'))})"],
        ["", ""],
        ["Income Received (12m)",     _gbp(s.get("income_ttm"))],
        ["Projected Annual Income",   _gbp(s.get("projected_annual_income"))],
        ["", ""],
        ["Total Positions",           str(s.get("positions_count", 0))],
    ]
    cover_ts = TableStyle([
        ("BACKGROUND",    (0,0), (0,-1), _PDF_SURF2),
        ("BACKGROUND",    (1,0), (1,-1), _PDF_SURF),
        ("TEXTCOLOR",     (0,0), (0,-1), _PDF_MUTED),
        ("TEXTCOLOR",     (1,0), (1,-1), _PDF_ACCENT),
        ("FONTNAME",      (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME",      (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 10),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("GRID",          (0,0), (-1,-1), 0.3, _PDF_BORDER),
        ("FONTNAME",      (0,2), (1,2), "Helvetica-Bold"),
        ("FONTSIZE",      (1,2), (1,2), 13),
        ("BACKGROUND",    (0,2), (1,2), _PDF_SURF),
        ("TEXTCOLOR",     (1,4), (1,4), _PDF_POS if cg >= 0 else _PDF_NEG),
        ("TEXTCOLOR",     (1,6), (1,7), _PDF_POS),
        ("BACKGROUND",    (0,5), (1,5), _PDF_BG),
        ("BACKGROUND",    (0,8), (1,8), _PDF_BG),
    ])
    W = doc.width
    story.append(Table(cover_data, colWidths=[W*0.45, W*0.55], style=cover_ts))
    story.append(Spacer(1, 8*mm))

    # AI Executive Summary
    story.append(Paragraph("Executive Summary (AI Generated)", h3))
    ai_text = await _fetch_ai_summary(d)
    for paragraph in ai_text.split('\n'):
        if paragraph.strip():
            story.append(Paragraph(paragraph.strip(), body_text))
    story.append(Spacer(1, 6*mm))

    # Asset breakdown
    story.append(Paragraph("Asset Class Breakdown", h3))
    at_data = [["Asset Class", "Positions", "Book Cost", "Market Value"]]
    for at, info in s.get("by_asset_type", {}).items():
        at_data.append([
            at.replace("_"," ").title(),
            str(info["count"]),
            _gbp(info["book_cost"]),
            _gbp(info["value"]),
        ])
    at_ts = _pdf_table_style()
    at_ts.add("ALIGN", (1,0), (3,-1), "RIGHT")
    story.append(Table(at_data, colWidths=[W*0.35, W*0.15, W*0.25, W*0.25], style=at_ts))

    # ── Holdings ───────────────────────────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Holdings", h2))

    hdr = ["Name", "Type", "Units", "Book/Unit", "Value", "PnL £", "PnL %", "Div Yield", "Source"]
    h_data = [hdr]
    for p in d["positions"]:
        pnl = p.get("unrealised_pnl"); pnl_pct = p.get("unrealised_pnl_pct")
        dy = p.get("annual_yield")
        h_data.append([
            p.get("name","")[:28],
            (p.get("asset_type") or "").replace("_"," ").title()[:10],
            _num(p.get("units"), 2),
            _gbp(p.get("book_cost_per_unit")),
            _gbp(p.get("current_value")),
            _gbp(pnl),
            _pct(pnl_pct),
            f"{dy*100:.2f}%" if dy is not None else "—",
            p.get("last_price_source") or "—",
        ])

    h_ts = _pdf_table_style()
    for row_i, p in enumerate(d["positions"], start=1):
        pnl = p.get("unrealised_pnl") or 0
        c = _PDF_POS if pnl >= 0 else _PDF_NEG
        h_ts.add("TEXTCOLOR", (5, row_i), (6, row_i), c)
    h_ts.add("ALIGN", (2,0), (-1,-1), "RIGHT")

    cws = [W*0.22, W*0.10, W*0.07, W*0.09, W*0.10, W*0.10, W*0.08, W*0.09, W*0.09]
    story.append(Table(h_data, colWidths=cws, style=h_ts, repeatRows=1))

    # ── Income Calendar ────────────────────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("12-Month Income Calendar", h2))

    cal = d["calendar"]
    grand = 0.0
    cws = [W*0.2, W*0.45, W*0.17, W*0.18]

    for ym in sorted(cal):
        lbl = datetime.strptime(ym+"-01","%Y-%m-%d").strftime("%B %Y")
        payments = sorted(cal[ym], key=lambda x: -x["projected_total"])
        mtotal = sum(p["projected_total"] for p in payments); grand += mtotal

        mhdr_ts = TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), _PDF_SURF2),
            ("TEXTCOLOR",    (0,0), (-1,0), _PDF_TEXT),
            ("TEXTCOLOR",    (3,0), (3,0),  _PDF_POS),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,0), 8),
            ("TOPPADDING",   (0,0), (-1,0), 4),
            ("BOTTOMPADDING",(0,0), (-1,0), 4),
            ("LEFTPADDING",  (0,0), (-1,0), 5),
            ("RIGHTPADDING", (0,0), (-1,0), 5),
            ("ALIGN",        (3,0), (3,0),  "RIGHT"),
            ("GRID",         (0,0), (-1,0), 0.3, _PDF_BORDER),
        ])
        story.append(Table([[lbl, "", "", _gbp(mtotal)]], colWidths=cws, style=mhdr_ts))

        det_data = [["", p["position_name"][:35], p["frequency_label"], _gbp(p["projected_total"])] for p in payments]
        det_ts = _pdf_table_style()
        det_ts.add("ALIGN", (3,0), (3,-1), "RIGHT")
        det_ts.add("TEXTCOLOR", (3,0), (3,-1), _PDF_POS)
        story.append(Table(det_data, colWidths=cws, style=det_ts))

    gt_ts = TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), _PDF_BG),
        ("TEXTCOLOR",    (0,0), (-1,0), _PDF_TEXT),
        ("TEXTCOLOR",    (3,0), (3,0),  _PDF_POS),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0), 9),
        ("TOPPADDING",   (0,0), (-1,0), 5),
        ("BOTTOMPADDING",(0,0), (-1,0), 5),
        ("LEFTPADDING",  (0,0), (-1,0), 5),
        ("RIGHTPADDING", (0,0), (-1,0), 5),
        ("ALIGN",        (3,0), (3,0),  "RIGHT"),
        ("GRID",         (0,0), (-1,0), 0.3, _PDF_BORDER),
    ])
    story.append(Table([["12-Month Total", "", "", _gbp(grand)]], colWidths=cws, style=gt_ts))

    # ── Received Dividends ─────────────────────────────────────────────────────
    if d["received"]:
        story.append(PageBreak())
        story.append(Paragraph("Received Dividends", h2))
        rec_data = [["Position", "Pay Date", "Amount", "Currency", "Notes"]]
        total = 0.0
        for r in d["received"]:
            rec_data.append([
                r.get("position_name","")[:30],
                r.get("pay_date",""),
                _gbp(r.get("amount")),
                r.get("currency",""),
                r.get("notes","") or "",
            ])
            total += r.get("amount") or 0
        rec_data.append(["Total Received", "", _gbp(total), "", ""])
        rec_ts = _pdf_table_style()
        rec_ts.add("ALIGN", (2,0), (2,-1), "RIGHT")
        last = len(rec_data)-1
        rec_ts.add("FONTNAME", (0,last), (-1,last), "Helvetica-Bold")
        rec_ts.add("BACKGROUND", (0,last), (-1,last), _PDF_SURF)
        rec_ts.add("TEXTCOLOR", (2,last), (2,last), _PDF_POS)
        story.append(Table(rec_data, colWidths=[W*0.32, W*0.15, W*0.15, W*0.10, W*0.28], style=rec_ts, repeatRows=1))

    def _bg(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(_PDF_BG)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        canvas.restoreState()

    doc.build(story, onFirstPage=_bg, onLaterPages=_bg)
    return buf.getvalue()


@router.get("/report.pdf")
async def export_pdf():
    d = _load_report_data()
    pdf_bytes = await _build_pdf_async(d)
    fname = f"portfolio_report_{d['iso_date']}.pdf"
    return StreamingResponse(io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})
